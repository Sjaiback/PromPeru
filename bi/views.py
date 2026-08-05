from collections import Counter
from datetime import datetime
from io import BytesIO
import hashlib, zipfile
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Avg, Count, DurationField, ExpressionWrapper, F, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.core.files.base import ContentFile
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from atencion.models import Atencion
from atencion.models import ArchivoAtenciones
from atencion.auditoria import registrar as auditar
from seguimiento.models import GestionAtencion


def es_bi(user):
    perfil = getattr(user, "perfil_asesor", None)
    return user.is_superuser or bool(perfil and perfil.activo and perfil.acceso_bi)


def puede_archivar(user):
    perfil = getattr(user, "perfil_asesor", None)
    return user.is_superuser or bool(perfil and perfil.activo and perfil.puede_archivar)


def filtradas(request):
    qs = Atencion.objects.select_related(
        "empresa__region", "empresa__sector", "responsable"
    )
    desde, hasta = request.GET.get("desde"), request.GET.get("hasta")
    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    if request.GET.get("responsable"):
        qs = qs.filter(responsable_id=request.GET["responsable"])
    if request.GET.get("sector"):
        qs = qs.filter(empresa__sector_id=request.GET["sector"])
    return qs


def series(qs, field, label):
    data = list(qs.values(field).annotate(total=Count("id")).order_by("-total")[:15])
    maximum = max([x["total"] for x in data], default=1)
    return [
        {
            "label": x[field] or "Sin dato",
            "total": x["total"],
            "pct": round(x["total"] * 100 / maximum),
        }
        for x in data
    ]


@login_required
@user_passes_test(es_bi)
def dashboard(request):
    from atencion.models import Responsable, Sector

    qs = filtradas(request)
    total = qs.count()
    gestiones = GestionAtencion.objects.filter(atencion__in=qs)
    estados = list(
        gestiones.values("estado__nombre", "estado__color")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    cerradas = gestiones.filter(estado__es_cerrado=True, resuelta__isnull=False)
    promedio = cerradas.annotate(
        duracion=ExpressionWrapper(
            F("resuelta") - F("iniciada"), output_field=DurationField()
        )
    ).aggregate(v=Avg("duracion"))["v"]
    recurrentes = (
        qs.values("empresa__id", "empresa__nombre", "empresa__numero_documento")
        .annotate(total=Count("id"))
        .filter(total__gt=1)
        .order_by("-total")[:10]
    )
    ratings = (
        qs.filter(empresa__rating__isnull=False)
        .values(
            "empresa__nombre",
            "empresa__numero_documento",
            "empresa__rating__total",
            "empresa__rating__clasificacion",
        )
        .distinct()
        .order_by("-empresa__rating__total")[:15]
    )
    context = {
        "total": total,
        "estados": estados,
        "promedio_horas": round(promedio.total_seconds() / 3600, 1) if promedio else 0,
        "atendidos": round(cerradas.count() * 100 / total, 1) if total else 0,
        "canales": series(qs, "tipo_atencion", "Canal"),
        "responsables_data": series(qs, "responsable__nombre", "Responsable"),
        "regiones_data": series(qs, "empresa__region__nombre", "Región"),
        "sectores_data": series(qs, "empresa__sector__nombre", "Sector"),
        "usuarios_data": series(qs, "empresa__tipo_usuario", "Tipo"),
        "recurrentes": recurrentes,
        "ratings": ratings,
        "responsables": Responsable.objects.filter(activo=True),
        "sectores": Sector.objects.filter(activo=True),
        "filtros": request.GET,
    }
    return render(request, "bi/dashboard.html", context)


@login_required
@user_passes_test(es_bi)
def exportar_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Atenciones"
    ws.append(
        [
            "Fecha",
            "Documento",
            "Empresa",
            "Canal",
            "Responsable",
            "Sector",
            "Región",
            "Tipo usuario",
            "Consulta",
            "Estado",
            "Rating",
        ]
    )
    for a in filtradas(request):
        gestion = getattr(a, "gestion", None)
        rating = getattr(a.empresa, "rating", None)
        ws.append(
            [
                a.fecha,
                a.empresa.numero_documento,
                a.empresa.nombre,
                a.tipo_atencion,
                a.responsable.nombre,
                a.empresa.sector.nombre,
                a.empresa.region.nombre,
                a.empresa.tipo_usuario,
                a.tema_consulta,
                gestion.estado.nombre if gestion else "",
                float(rating.total) if rating else "",
            ]
        )
    for index, col in enumerate(ws.columns, start=1):
        ws.column_dimensions[get_column_letter(index)].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 45
        )
    out = BytesIO()
    wb.save(out)
    response = HttpResponse(
        out.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_atenciones.xlsx"'
    return response


@login_required
@user_passes_test(es_bi)
def exportar_pdf(request):
    out = BytesIO()
    pdf = canvas.Canvas(out, pagesize=A4)
    width, height = A4
    pdf.setTitle("Reporte de atenciones PROMPERÚ")
    y = height - 50
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(45, y, "PROMPERÚ · Reporte de atenciones")
    y -= 28
    pdf.setFont("Helvetica", 9)
    for a in filtradas(request):
        line = f"{a.fecha:%d/%m/%Y} | {a.empresa.numero_documento} | {a.empresa.nombre[:40]} | {a.tipo_atencion} | {a.responsable.nombre[:28]}"
        pdf.drawString(45, y, line)
        y -= 15
        if y < 45:
            pdf.showPage()
            pdf.setFont("Helvetica", 9)
            y = height - 45
    pdf.save()
    response = HttpResponse(out.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_atenciones.pdf"'
    return response


def _workbook_atenciones(qs, titulo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Atenciones"
    ws.append([titulo])
    ws.merge_cells("A1:R1")
    ws.append(
        [
            "ID",
            "Fecha",
            "Canal",
            "Responsable",
            "Documento",
            "Empresa/persona",
            "Contacto",
            "Cargo",
            "Teléfono",
            "Email",
            "Sector",
            "Región",
            "Tipo usuario",
            "Producto/interés",
            "Consulta",
            "Estado",
            "Seguimientos",
            "Rating",
        ]
    )
    for a in qs.select_related(
        "empresa__sector", "empresa__region", "responsable"
    ).prefetch_related("gestion__seguimientos"):
        gestion = getattr(a, "gestion", None)
        rating = getattr(a.empresa, "rating", None)
        logs = (
            " | ".join(
                f"{x.fecha_hora:%d/%m/%Y}: {x.detalle}"
                for x in gestion.seguimientos.all()
            )
            if gestion
            else ""
        )
        ws.append(
            [
                a.pk,
                a.fecha,
                a.tipo_atencion,
                a.responsable.nombre,
                a.empresa.numero_documento,
                a.empresa.nombre,
                a.empresa.nombres_apellidos,
                a.empresa.cargo,
                a.empresa.telefono,
                a.empresa.email,
                a.empresa.sector.nombre,
                a.empresa.region.nombre,
                a.empresa.tipo_usuario,
                a.empresa.oferta_producto_servicio,
                a.tema_consulta,
                gestion.estado.nombre if gestion else "",
                logs,
                float(rating.total) if rating else "",
            ]
        )
    for cell in ws[2]:
        cell.font = cell.font.copy(bold=True)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:R{max(ws.max_row, 2)}"
    for index, col in enumerate(ws.columns, start=1):
        ws.column_dimensions[get_column_letter(index)].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 45
        )
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@login_required
@user_passes_test(puede_archivar)
def archivos(request):
    if request.method == "POST":
        desde, hasta = request.POST.get("desde"), request.POST.get("hasta")
        if not desde or not hasta:
            messages.error(request, "Indica ambas fechas.")
        else:
            qs = Atencion.objects.filter(
                fecha__range=[desde, hasta], anulada=False
            ).order_by("responsable__nombre", "fecha")
            if not qs.exists():
                messages.error(request, "No hay atenciones en ese periodo.")
            else:
                bundle = BytesIO()
                with zipfile.ZipFile(bundle, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr(
                        "00_GLOBAL_atenciones.xlsx",
                        _workbook_atenciones(
                            qs, "PROMPERÚ · Archivo global de atenciones"
                        ),
                    )
                    for responsable_id, nombre in qs.values_list(
                        "responsable_id", "responsable__nombre"
                    ).distinct():
                        safe = "".join(c if c.isalnum() else "_" for c in nombre).strip(
                            "_"
                        )
                        zf.writestr(
                            f"asesores/{safe}.xlsx",
                            _workbook_atenciones(
                                qs.filter(responsable_id=responsable_id),
                                f"PROMPERÚ · {nombre}",
                            ),
                        )
                payload = bundle.getvalue()
                digest = hashlib.sha256(payload).hexdigest()
                archivo = ArchivoAtenciones(
                    desde=desde,
                    hasta=hasta,
                    generado_por=request.user,
                    checksum_sha256=digest,
                    total_atenciones=qs.count(),
                )
                archivo.archivo.save(
                    f"atenciones_{desde}_{hasta}.zip", ContentFile(payload), save=True
                )
                auditar(
                    request,
                    "archivar",
                    archivo,
                    descripcion=f"Archivo generado con {qs.count()} atenciones",
                )
                messages.success(
                    request,
                    "Archivo ZIP generado. Descárgalo y verifícalo antes de depurar.",
                )
                return redirect("bi:archivos")
    return render(
        request, "bi/archivos.html", {"archivos": ArchivoAtenciones.objects.all()}
    )


@login_required
@user_passes_test(puede_archivar)
def descargar_archivo(request, pk):
    obj = get_object_or_404(ArchivoAtenciones, pk=pk)
    if not obj.descargado_en:
        obj.descargado_en = timezone.now()
        obj.save(update_fields=["descargado_en"])
    response = HttpResponse(
        obj.archivo.open("rb").read(), content_type="application/zip"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{obj.archivo.name.split("/")[-1]}"'
    )
    return response


@login_required
@user_passes_test(puede_archivar)
def depurar_archivo(request, pk):
    obj = get_object_or_404(ArchivoAtenciones, pk=pk, depurado=False)
    if not obj.descargado_en:
        messages.error(
            request, "Primero debes descargar el archivo ZIP antes de depurar."
        )
        return redirect("bi:archivos")
    if request.method == "POST" and request.POST.get("confirmacion") == "DEPURAR":
        qs = Atencion.objects.filter(fecha__range=[obj.desde, obj.hasta])
        ids = list(qs.values_list("pk", flat=True))
        total = len(ids)
        qs.delete()
        obj.depurado = True
        obj.depurado_en = timezone.now()
        obj.save(update_fields=["depurado", "depurado_en"])
        auditar(
            request,
            "depurar",
            obj,
            antes={"atenciones": ids},
            despues={"atenciones_eliminadas": total},
            descripcion="Depuración posterior a archivo; empresas y asesores conservados",
        )
        messages.success(
            request,
            f"Se liberaron {total} atenciones. Empresas, personas y asesores se conservaron.",
        )
        return redirect("bi:archivos")
    return render(request, "bi/depurar.html", {"archivo": obj})
