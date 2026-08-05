import re
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.core.management.base import BaseCommand, CommandError
from django.utils.text import slugify
from rating.models import CategoriaRating, CriterioRating


def limpio(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def es_columna_puntaje(texto, muestra):
    t = limpio(texto).lower()
    patrones = [
        r"^\d+(?:[.,]\d+)?$",
        r"^si\s*=",
        r"^>=",
        r"^exp\s*=",
        r"^estudio o plan aprobado",
        r"^logística si",
    ]
    if any(re.search(p, t) for p in patrones):
        return True
    return isinstance(muestra, (int, float)) and (not texto or len(t) < 18)


class Command(BaseCommand):
    help = "Convierte los encabezados multinivel del Excel histórico en criterios editables"

    def add_arguments(self, parser):
        parser.add_argument("archivo")

    def handle(self, *args, **opts):
        try:
            ws = load_workbook(opts["archivo"], data_only=False).active
        except Exception as exc:
            raise CommandError(str(exc))
        merged = {}
        for rng in ws.merged_cells.ranges:
            value = ws.cell(rng.min_row, rng.min_col).value
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    merged[(row, col)] = value
        created = 0
        order = 0
        for col in range(1, ws.max_column + 1):
            category = (
                limpio(merged.get((5, col), ws.cell(5, col).value)) or "Otros criterios"
            )
            group = limpio(merged.get((6, col), ws.cell(6, col).value))
            question = limpio(ws.cell(7, col).value) or group
            sample = ws.cell(8, col).value
            if col <= 14 or not question or es_columna_puntaje(question, sample):
                continue
            if question in ("#REF!", "RATING") or category == "TOTAL":
                continue
            label = (
                question if not group or group == question else f"{group} — {question}"
            )
            cat, _ = CategoriaRating.objects.get_or_create(
                nombre=category[:180],
                defaults={
                    "slug": slugify(category)[:180] or f"categoria-{col}",
                    "orden": col,
                },
            )
            if (
                CriterioRating.objects.exclude(origen_excel=f"columna:{col}")
                .filter(categoria=cat, nombre=label[:255])
                .exists()
            ):
                label = f"{label} ({get_column_letter(col)})"
            code_base = slugify(f"{category}-{label}")[:220] or f"criterio-{col}"
            code = code_base
            suffix = 2
            while (
                CriterioRating.objects.exclude(origen_excel=f"columna:{col}")
                .filter(codigo=code)
                .exists()
            ):
                code = f"{code_base[:215]}-{suffix}"
                suffix += 1
            tipo = (
                "booleano"
                if re.search(r"si\s*/?\s*no|sí\s*/?\s*no", label, re.I)
                else (
                    "numero"
                    if re.search(
                        r"nº|n°|ventas|capacidad|tiempo|años|meses|cuántos", label, re.I
                    )
                    else "texto"
                )
            )
            obj, was_created = CriterioRating.objects.update_or_create(
                origen_excel=f"columna:{col}",
                defaults={
                    "categoria": cat,
                    "nombre": label[:255],
                    "codigo": code,
                    "tipo_dato": tipo,
                    "orden": order,
                    "ayuda": limpio(group if group != question else ""),
                    "opciones": ["Sí", "No"] if tipo == "booleano" else [],
                    "activo": True,
                },
            )
            created += int(was_created)
            order += 1
        self.stdout.write(
            self.style.SUCCESS(f"Criterios preparados: {order} ({created} nuevos).")
        )
