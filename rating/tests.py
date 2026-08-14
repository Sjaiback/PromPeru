from io import BytesIO
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from openpyxl import Workbook
from .views import COLUMNAS_CLIENTES, leer_excel
from .ficha import puntuar
from .models import EvaluacionVisita, PerfilEvaluacionEmpresa
from atencion.models import Atencion, Empresa, PerfilAsesor, Region, Responsable, Sector


class ExcelParserTests(SimpleTestCase):
    def test_lee_encabezados_multinivel_y_filas(self):
        wb = Workbook()
        ws = wb.active
        ws.merge_cells("A5:A7")
        ws["A5"] = "RUC"
        ws.merge_cells("B5:C5")
        ws["B5"] = "Información Básica"
        ws["B6"] = "Razón Social"
        ws["C6"] = "Sector"
        ws["B7"] = "Valor"
        ws["C7"] = "Valor"
        ws.append([])
        ws["A8"] = "20123456789"
        ws["B8"] = "Empresa"
        ws["C8"] = "Servicios"
        out = BytesIO()
        wb.save(out)
        headers, rows = leer_excel(out.getvalue(), 5, 3)
        self.assertIn("RUC", headers[0])
        self.assertIn("Razón Social", headers[1])
        self.assertEqual(rows[0][0], "20123456789")


class PuntajeFichaTests(SimpleTestCase):
    def test_calcula_maximos_confirmados_de_secciones_principales(self):
        respuestas = {
            "vigencia_18": "si", "correo_corporativo": "si", "web": "si",
            "redes": "si", "catalogo_es": "si", "catalogo_en": "si",
            "deudas_promperu": "no", "sentinel": "1.99",
            "web_responsiva": "si", "web_usabilidad": "si", "web_actualizada": "si",
            "web_idioma": "si", "facebook": "si", "instagram": "si",
            "publicidad_online": "si", "plan_contenido": "si", "catalogo_virtual": "si",
            "manual_marca": "si", "video_corporativo": "si",
            "plan_negocio": "si", "estructura_costos": "si",
            "tipo_estructura_costos": "exportacion", "estudio_impacto": "si",
            "proyecto_innovacion": "si", "equipo_investigacion": "si",
            "propuesta_valor": "si", "nivel_ingles": "avanzado", "incoterm": "fob",
            "responsable_comercio_exterior": "si",
        }
        puntos, _ = puntuar(respuestas)
        self.assertEqual(puntos["obligatorios"], 8)
        self.assertEqual(puntos["digitalizacion"], 4)
        self.assertEqual(puntos["financiera"], 2.5)
        self.assertEqual(puntos["gobierno"], 3.75)


class FichaPorVisitaTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user("evaluador", password="segura-123")
        region = Region.objects.create(nombre="Junín")
        sector = Sector.objects.create(nombre="Servicios")
        responsable = Responsable.objects.create(nombre="Coordinador - Prueba")
        PerfilAsesor.objects.create(usuario=self.user, responsable=responsable, rol="asesor")
        self.empresa = Empresa.objects.create(
            tipo_documento="DNI", numero_documento="12345678", nombre="Persona prueba",
            nombres_apellidos="Ana Pérez", cargo="Gerente", tipo_usuario="Exportador",
            telefono="999999999", email="ana@example.test", sector=sector, region=region,
            oferta_producto_servicio="Servicios",
        )
        self.atencion = Atencion.objects.create(empresa=self.empresa, responsable=responsable, tipo_atencion="Presencial")
        self.client.force_login(self.user)

    def test_guarda_ficha_y_reutiliza_datos_de_empresa(self):
        url = reverse("rating:evaluar", args=[self.empresa.pk])
        self.assertEqual(self.client.get(url).status_code, 200)
        response = self.client.post(url, {"vigencia_18": "si", "web": "si", "web_url": "https://example.test"})
        self.assertRedirects(response, url)
        evaluacion = EvaluacionVisita.objects.get(empresa=self.empresa)
        self.assertEqual(evaluacion.respuestas["web_url"], "https://example.test")
        self.assertEqual(PerfilEvaluacionEmpresa.objects.get(empresa=self.empresa).datos["web"], "si")
        self.assertContains(self.client.get(url), "https://example.test")
        segunda = Atencion.objects.create(
            empresa=self.empresa,
            responsable=self.atencion.responsable,
            tipo_atencion="WhatsApp",
        )
        segunda_url = reverse("rating:evaluar", args=[self.empresa.pk])
        self.assertContains(self.client.get(segunda_url), "https://example.test")
        self.client.post(segunda_url, {"web": "si", "web_url": "https://example.test"})
        self.assertEqual(EvaluacionVisita.objects.filter(empresa=self.empresa).count(), 1)
        evaluacion.refresh_from_db()
        export = self.client.get(reverse("rating:exportar_evaluacion", args=[evaluacion.pk]))
        self.assertEqual(export.status_code, 200)
        self.assertIn("evaluacion_", export["Content-Disposition"])

    def test_evalua_empresa_sin_atencion_previa(self):
        empresa = Empresa.objects.create(
            tipo_documento="RUC", numero_documento="20999999991",
            nombre="Empresa sin consulta", tipo_usuario="Exportador",
            telefono="999888111", email="sinconsulta@example.test",
            sector=self.empresa.sector, region=self.empresa.region,
            oferta_producto_servicio="Café",
        )
        url = reverse("rating:evaluar", args=[empresa.pk])
        self.assertEqual(self.client.get(url).status_code, 200)
        self.client.post(url, {"vigencia_18": "si"})
        self.assertTrue(EvaluacionVisita.objects.filter(empresa=empresa).exists())

    def test_importa_solo_clientes_sin_crear_atenciones(self):
        self.user.is_staff = True
        self.user.save(update_fields=["is_staff"])
        wb = Workbook()
        ws = wb.active
        headers = sorted(COLUMNAS_CLIENTES)
        ws.append(headers)
        row = {header: "" for header in headers}
        row.update(
            {
                "SECTOR": "Servicios",
                "LÍNEA": "",
                "PRODUCTO": "Software",
                "REGIÓN": "Junín",
                "TIPO DE DOCUMENTO": "RUC",
                "N° DEL DOCUMENTO": "20999999992",
                "EMPRESA / INSTITUCIÓN / PERSONA NATURAL": "Cliente importado",
                "TIPO DE USUARIO": "Exportador",
                "TIPO DE PERSONERÍA": "Persona Jurídica",
                "NOMBRE Y APELLIDO": "Rosa Quispe",
                "CARGO": "Gerente",
                "TELEFONO/CELULAR": "999111999",
                "E-MAIL": "rosa@example.test",
            }
        )
        ws.append([row[header] for header in headers])
        out = BytesIO()
        wb.save(out)
        response = self.client.post(
            reverse("rating:intercambiar"),
            {
                "tipo_importacion": "clientes",
                "archivo": SimpleUploadedFile(
                    "clientes.xlsx", out.getvalue(),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        self.assertRedirects(response, reverse("rating:intercambiar"))
        empresa = Empresa.objects.get(numero_documento="20999999992")
        self.assertEqual(empresa.nombre, "Cliente importado")
        self.assertFalse(empresa.atenciones.exists())
