import io, re, unicodedata
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
import re
from django.utils.text import slugify
from django.utils import timezone
from atencion.models import (
    Atencion,
    Empresa,
    Region,
    Responsable,
    RespaldoLimpieza,
    Sector,
)
from atencion.permisos import es_cuenta_sistemas
from bi.supabase_storage import descargar_respaldo, subir_respaldo
from seguimiento.models import EstadoAtencion, GestionAtencion
from openpyxl import Workbook
from .forms import ImportarExcelForm, ImportarDatosForm
from .forms import EvaluacionDinamicaForm
from .models import (
    CategoriaRating,
    CriterioRating,
    EmpresaRating,
    ImportacionRating,
    MapeoColumna,
    ValorCriterio,
    EvaluacionVisita,
    PerfilEvaluacionEmpresa,
)
from atencion.auditoria import registrar as auditar
from .ficha import SECCIONES, preparar_ficha, puntuar


def norm(value):
    value = (
        unicodedata.normalize("NFKD", str(value or ""))
        .encode("ascii", "ignore")
        .decode()
        .lower()
    )
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


COLUMNAS_INTERCAMBIO = [
    "N°", "FECHA", "TIPO DE ATENCIÓN", "RESPONSABLE", "SECTOR", "LÍNEA",
    "PRODUCTO", "REGIÓN", "TIPO DE DOCUMENTO", "N° DEL DOCUMENTO",
    "EMPRESA / INSTITUCIÓN / PERSONA NATURAL", "TIPO DE USUARIO",
    "TIPO DE PERSONERÍA", "NOMBRE Y APELLIDO", "CARGO", "TELEFONO/CELULAR",
    "E-MAIL", "TEMA DE CONSULTA", "DETALLAR CONSULTA", "ACCIÓN REALIZADA",
    "ESTADO DE LA ATENCIÓN", "SEGUIMIENTO", "OBSERVACIONES",
]

COLUMNAS_CLIENTES = {
    "SECTOR", "LÍNEA", "PRODUCTO", "REGIÓN", "TIPO DE DOCUMENTO",
    "N° DEL DOCUMENTO", "EMPRESA / INSTITUCIÓN / PERSONA NATURAL",
    "TIPO DE USUARIO", "TIPO DE PERSONERÍA", "NOMBRE Y APELLIDO", "CARGO",
    "TELEFONO/CELULAR", "E-MAIL",
}


def _puede_intercambiar(user):
    perfil = getattr(user, "perfil_asesor", None)
    return user.is_superuser or user.is_staff or bool(
        perfil
        and perfil.activo
        and (perfil.puede_archivar or perfil.rol in {"admin", "coordinador"})
    )


def _texto(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _fecha_hora(value):
    if isinstance(value, datetime):
        return value.date(), value.time().replace(microsecond=0)
    if isinstance(value, date):
        return value, time(0, 0)
    raw = _texto(value)
    for formato in (
        "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"
    ):
        try:
            parsed = datetime.strptime(raw, formato)
            return parsed.date(), parsed.time()
        except ValueError:
            continue
    return timezone.localdate(), timezone.localtime().time().replace(microsecond=0)


def _catalogo(modelo, nombre, por_defecto):
    valor = _texto(nombre) or por_defecto
    existente = modelo.objects.filter(nombre__iexact=valor).first()
    return existente or modelo.objects.create(nombre=valor)


def _empresa_desde_fila(fila):
    documento = _texto(fila.get("N° DEL DOCUMENTO"))
    if not documento:
        raise ValueError("La fila no tiene N° DEL DOCUMENTO")
    tipo_documento = _texto(fila.get("TIPO DE DOCUMENTO")) or (
        "RUC" if len(documento) == 11 else "DNI"
    )
    sector = _catalogo(Sector, fila.get("SECTOR"), "Otros")
    region = _catalogo(Region, fila.get("REGIÓN"), "Sin especificar")
    defaults = {
        "nombre": _texto(fila.get("EMPRESA / INSTITUCIÓN / PERSONA NATURAL")) or documento,
        "tipo_usuario": _texto(fila.get("TIPO DE USUARIO")) or "Empresa con Potencial Exportador",
        "tipo_personeria": _texto(fila.get("TIPO DE PERSONERÍA")),
        "nombres_apellidos": _texto(fila.get("NOMBRE Y APELLIDO")),
        "cargo": _texto(fila.get("CARGO")),
        "telefono": _texto(fila.get("TELEFONO/CELULAR")),
        "email": _texto(fila.get("E-MAIL")),
        "sector": sector,
        "region": region,
        "oferta_producto_servicio": _texto(fila.get("PRODUCTO")) or "Otros",
        "activa": True,
    }
    empresa, creada = Empresa.objects.update_or_create(
        tipo_documento=tipo_documento,
        numero_documento=documento,
        defaults=defaults,
    )
    rating, _ = EmpresaRating.objects.get_or_create(empresa=empresa)
    rating.linea = _texto(fila.get("LÍNEA"))
    rating.producto = defaults["oferta_producto_servicio"]
    rating.save(update_fields=["linea", "producto", "actualizado"])
    return empresa, creada


@login_required
@user_passes_test(_puede_intercambiar)
def intercambiar_datos(request):
    form = ImportarDatosForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        upload = form.cleaned_data["archivo"]
        if not upload.name.lower().endswith(".xlsx"):
            form.add_error("archivo", "Solo se admiten archivos .xlsx.")
        else:
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            canonicos = {norm(columna): columna for columna in COLUMNAS_INTERCAMBIO}
            encabezados = [
                canonicos.get(norm(cell.value), _texto(cell.value)) for cell in sheet[1]
            ]
            requeridas = (
                set(COLUMNAS_INTERCAMBIO)
                if form.cleaned_data["tipo_importacion"] == "atenciones"
                else COLUMNAS_CLIENTES
            )
            faltantes = [col for col in requeridas if col not in encabezados]
            if faltantes:
                form.add_error(
                    "archivo",
                    "Faltan columnas obligatorias: " + ", ".join(sorted(faltantes)),
                )
            else:
                indices = {header: idx for idx, header in enumerate(encabezados)}
                creadas = actualizadas = fallidas = 0
                errores = []
                log = ImportacionRating.objects.create(
                    archivo_nombre=upload.name,
                    usuario=request.user,
                    filas_detectadas=max(sheet.max_row - 1, 0),
                )
                for numero_fila, values in enumerate(
                    sheet.iter_rows(min_row=2, values_only=True), start=2
                ):
                    if not any(value not in (None, "") for value in values):
                        continue
                    fila = {
                        col: values[idx] if idx < len(values) else ""
                        for col, idx in indices.items()
                    }
                    try:
                        with transaction.atomic():
                            empresa, creada = _empresa_desde_fila(fila)
                            if form.cleaned_data["tipo_importacion"] == "atenciones":
                                fecha, hora = _fecha_hora(fila.get("FECHA"))
                                responsable = _catalogo(
                                    Responsable, fila.get("RESPONSABLE"), "Sin asignar"
                                )
                                id_origen = _texto(fila.get("N°"))
                                atencion = (
                                    Atencion.objects.filter(pk=int(id_origen), empresa=empresa).first()
                                    if id_origen.isdigit()
                                    else None
                                )
                                datos_atencion = {
                                    "fecha": fecha,
                                    "hora": hora,
                                    "tipo_atencion": _texto(fila.get("TIPO DE ATENCIÓN")) or "Presencial",
                                    "responsable": responsable,
                                    "empresa": empresa,
                                    "tema_consulta": _texto(fila.get("TEMA DE CONSULTA")),
                                    "detalle_consulta": _texto(fila.get("DETALLAR CONSULTA")),
                                    "origen": "importado",
                                    "registrado_por": request.user,
                                }
                                if atencion:
                                    for campo, valor in datos_atencion.items():
                                        setattr(atencion, campo, valor)
                                    atencion.save()
                                else:
                                    atencion = Atencion.objects.create(**datos_atencion)
                                seguimiento_txt = _texto(fila.get("SEGUIMIENTO")).lower()
                                finalizado = seguimiento_txt == "finalizado"
                                estado_legacy = _catalogo(
                                    EstadoAtencion,
                                    "Finalizado" if finalizado else "En proceso",
                                    "En proceso",
                                )
                                estado_legacy.es_cerrado = finalizado
                                estado_legacy.save(update_fields=["es_cerrado"])
                                estado_txt = _texto(fila.get("ESTADO DE LA ATENCIÓN")).lower()
                                mapa_estado = {
                                    "atendido": "atendido",
                                    "atendido y derivado": "atendido_derivado",
                                    "sin atender": "sin_atender",
                                }
                                GestionAtencion.objects.update_or_create(
                                    atencion=atencion,
                                    defaults={
                                        "estado": estado_legacy,
                                        "accion_realizada": _texto(fila.get("ACCIÓN REALIZADA")),
                                        "estado_atencion": mapa_estado.get(estado_txt, "sin_atender"),
                                        "estado_seguimiento": "finalizado" if finalizado else "en_proceso",
                                        "observaciones": _texto(fila.get("OBSERVACIONES")),
                                        "actualizado_por": request.user,
                                    },
                                )
                            creadas += int(creada)
                            actualizadas += int(not creada)
                    except Exception as exc:
                        fallidas += 1
                        errores.append(f"Fila {numero_fila}: {exc}")
                log.filas_importadas = creadas
                log.filas_actualizadas = actualizadas
                log.filas_fallidas = fallidas
                log.detalle = {
                    "errores": errores[:50],
                    "tipo": form.cleaned_data["tipo_importacion"],
                }
                log.save()
                messages.success(
                    request,
                    f"Importación terminada: {creadas} nuevos, {actualizadas} actualizados y {fallidas} fallidos.",
                )
                return redirect("rating:intercambiar")
    return render(
        request,
        "rating/intercambiar.html",
        {
            "form": form,
            "logs": ImportacionRating.objects.all()[:10],
            "columnas": COLUMNAS_INTERCAMBIO,
        },
    )


@login_required
@user_passes_test(_puede_intercambiar)
def limpiar_datos_exportados(request):
    respaldo = request.session.get("ultima_exportacion_limpieza")
    if request.method == "POST":
        if not respaldo:
            messages.error(
                request,
                "Primero debes exportar y descargar el Excel desde esta pantalla.",
            )
            return redirect("rating:limpiar_datos")
        if request.POST.get("datos_exportados") != "1":
            messages.error(request, "Confirma que exportaste y guardaste los datos.")
        elif request.POST.get("confirmacion", "").strip().upper() != "LIMPIAR":
            messages.error(request, "Escribe LIMPIAR para confirmar la operación.")
        else:
            hasta = datetime.fromisoformat(respaldo["hasta"])
            qs = Atencion.objects.filter(anulada=False, creado__lte=hasta).order_by(
                "fecha", "hora", "pk"
            )
            ids_muestra = list(qs.values_list("pk", flat=True)[:200])
            total = qs.count()
            if not total:
                messages.info(request, "No hay atenciones exportadas para eliminar.")
                return redirect("rating:intercambiar")

            desde = qs.values_list("fecha", flat=True).first()
            fecha_limpieza = timezone.localdate()
            marca = timezone.localtime().strftime("%Y%m%d_%H%M%S")
            nombre = (
                f"atenciones_{desde:%Y-%m-%d}_a_"
                f"{fecha_limpieza:%Y-%m-%d}_{marca}.xlsx"
            )
            ruta_storage = (
                f"limpiezas/{fecha_limpieza:%Y/%m}/{nombre}"
            )
            from bi.views import _workbook_atenciones

            contenido = _workbook_atenciones(
                qs,
                "PROMPERÚ · Respaldo automático previo a limpieza",
            )
            try:
                storage = subir_respaldo(contenido, ruta_storage)
            except Exception as exc:
                messages.error(
                    request,
                    "No se eliminó ningún dato porque el respaldo privado no pudo "
                    f"guardarse en Supabase. Detalle: {exc}",
                )
                return redirect("rating:limpiar_datos")

            with transaction.atomic():
                respaldo_guardado = RespaldoLimpieza.objects.create(
                    bucket=storage["bucket"],
                    ruta_storage=storage["ruta"],
                    nombre_archivo=nombre,
                    desde=desde,
                    hasta=fecha_limpieza,
                    total_atenciones=total,
                    tamano_bytes=storage["tamano"],
                    checksum_sha256=storage["checksum"],
                    generado_por=request.user,
                )
                qs.delete()
                auditar(
                    request,
                    "depurar",
                    entidad="Atencion",
                    antes={
                        "total": total,
                        "exportado_hasta": respaldo["hasta"],
                        "ids_muestra": ids_muestra,
                    },
                    despues={
                        "atenciones_eliminadas": total,
                        "empresas_y_usuarios_conservados": True,
                        "respaldo_storage": respaldo_guardado.ruta_storage,
                    },
                    descripcion=(
                        f"Limpieza de {total} atenciones previamente exportadas; "
                        "empresas, evaluaciones y usuarios conservados"
                    ),
                )
            request.session.pop("ultima_exportacion_limpieza", None)
            messages.success(
                request,
                f"Se respaldaron y eliminaron {total} atenciones. Las empresas, "
                "evaluaciones y cuentas se conservaron.",
            )
            return redirect("rating:intercambiar")
    return render(
        request,
        "rating/limpiar_datos.html",
        {"respaldo": respaldo},
    )


@login_required
def respaldos_limpieza(request):
    if not es_cuenta_sistemas(request.user):
        return render(request, "errors/403.html", status=403)
    return render(
        request,
        "rating/respaldos.html",
        {"respaldos": RespaldoLimpieza.objects.select_related("generado_por")},
    )


@login_required
def descargar_respaldo_limpieza(request, pk):
    if not es_cuenta_sistemas(request.user):
        return render(request, "errors/403.html", status=403)
    respaldo = get_object_or_404(RespaldoLimpieza, pk=pk)
    try:
        contenido = descargar_respaldo(
            respaldo.bucket,
            respaldo.ruta_storage,
        )
    except Exception as exc:
        messages.error(request, f"No se pudo descargar el respaldo: {exc}")
        return redirect("rating:respaldos")
    response = HttpResponse(
        contenido,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{respaldo.nombre_archivo}"'
    )
    return response


ALIASES = {
    "ruc": "empresa.numero_documento",
    "razon social": "empresa.nombre",
    "razon_social": "empresa.nombre",
    "sector": "empresa.sector",
    "ubicacion": "empresa.region",
    "region": "empresa.region",
    "condicion sunat": "rating.condicion_sunat",
    "linea": "rating.linea",
    "producto": "rating.producto",
    "total": "rating.total",
    "rating": "rating.total",
}


def leer_excel(data, fila_inicio, cantidad):
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    ws = wb.active
    # Propaga valores de celdas combinadas para reconstruir encabezados multinivel.
    merged = {}
    for rng in ws.merged_cells.ranges:
        top = ws.cell(rng.min_row, rng.min_col).value
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merged[(row, col)] = top
    headers = []
    for col in range(1, ws.max_column + 1):
        parts = []
        for row in range(fila_inicio, fila_inicio + cantidad):
            value = merged.get((row, col), ws.cell(row, col).value)
            if value and str(value).strip() not in parts:
                parts.append(str(value).strip())
        headers.append(" / ".join(parts) or f"Columna {col}")
    rows = []
    for values in ws.iter_rows(min_row=fila_inicio + cantidad, values_only=True):
        if any(v not in (None, "") for v in values):
            rows.append(["" if v is None else str(v) for v in values])
    return headers, rows


def sugerir(header):
    n = norm(header)
    saved = MapeoColumna.objects.filter(encabezado_origen=n, activo=True).first()
    if saved:
        return saved.campo_destino
    for alias, target in sorted(ALIASES.items(), key=lambda x: -len(x[0])):
        if norm(alias) in n:
            return target
    return ""


@staff_member_required
def importar(request):
    form = ImportarExcelForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        upload = form.cleaned_data["archivo"]
        if not upload.name.lower().endswith(".xlsx"):
            form.add_error("archivo", "Solo se admiten archivos .xlsx.")
        else:
            headers, rows = leer_excel(
                upload.read(),
                form.cleaned_data["fila_inicio_encabezado"],
                form.cleaned_data["filas_encabezado"],
            )
            request.session["rating_preview"] = {
                "archivo": upload.name,
                "headers": headers,
                "rows": rows[:1000],
            }
            columnas = [
                {"indice": i, "encabezado": h, "sugerencia": sugerir(h)}
                for i, h in enumerate(headers)
            ]
            return render(
                request,
                "rating/preview.html",
                {
                    "headers": headers,
                    "columnas": columnas,
                    "rows": rows[:10],
                    "total": len(rows),
                },
            )
    return render(
        request,
        "rating/importar.html",
        {"form": form, "logs": ImportacionRating.objects.all()[:10]},
    )


@staff_member_required
@transaction.atomic
def confirmar(request):
    if request.method != "POST" or "rating_preview" not in request.session:
        return redirect("rating:importar")
    data = request.session["rating_preview"]
    headers = data["headers"]
    rows = data["rows"]
    mappings = [request.POST.get(f"map_{i}", "").strip() for i in range(len(headers))]
    duplicate_mode = request.POST.get("duplicados", "actualizar")
    log = ImportacionRating.objects.create(
        archivo_nombre=data["archivo"], usuario=request.user, filas_detectadas=len(rows)
    )
    errors = []
    for idx, row in enumerate(rows, start=1):
        record = {m: row[i] for i, m in enumerate(mappings) if m and i < len(row)}
        doc = re.sub(r"\.0$", "", record.get("empresa.numero_documento", "")).strip()
        if not doc:
            errors.append({"fila": idx, "error": "RUC/documento no identificado"})
            continue
        existing = Empresa.objects.filter(numero_documento=doc).first()
        if existing and duplicate_mode == "ignorar":
            log.filas_ignoradas += 1
            continue
        sector_name = record.get("empresa.sector", "Sin clasificar") or "Sin clasificar"
        region_name = record.get("empresa.region", "Lima") or "Lima"
        sector, _ = Sector.objects.get_or_create(nombre=sector_name)
        region, _ = Region.objects.get_or_create(nombre=region_name)
        defaults = {
            "nombre": record.get(
                "empresa.nombre", existing.nombre if existing else f"Empresa {doc}"
            ),
            "sector": sector,
            "region": region,
            "tipo_usuario": "Exportador",
            "telefono": existing.telefono if existing else "000000000",
            "email": existing.email if existing else "pendiente@example.local",
            "oferta_producto_servicio": record.get("rating.producto", "Por completar"),
        }
        empresa, created = Empresa.objects.update_or_create(
            numero_documento=doc,
            defaults={"tipo_documento": "RUC" if len(doc) == 11 else "DNI", **defaults},
        )
        rating, _ = EmpresaRating.objects.get_or_create(empresa=empresa)
        for field in ("condicion_sunat", "linea", "producto"):
            if f"rating.{field}" in record:
                setattr(rating, field, record[f"rating.{field}"])
        try:
            rating.total = Decimal(
                str(record.get("rating.total", rating.total)).replace(",", ".")
            )
        except InvalidOperation:
            pass
        rating.save()
        for destination, value in record.items():
            if not destination.startswith("criterio:"):
                continue
            code = destination.split(":", 1)[1]
            criterio = CriterioRating.objects.filter(codigo=code).first()
            if criterio:
                ValorCriterio.objects.update_or_create(
                    rating=rating,
                    criterio=criterio,
                    defaults={"valor_texto": str(value)},
                )
        for header, destination in zip(headers, mappings):
            if destination:
                MapeoColumna.objects.update_or_create(
                    encabezado_origen=norm(header),
                    defaults={"campo_destino": destination},
                )
        if created:
            log.filas_importadas += 1
        else:
            log.filas_actualizadas += 1
    log.filas_fallidas = len(errors)
    log.detalle = {"errores": errors}
    log.save()
    del request.session["rating_preview"]
    messages.success(
        request,
        f"Importación terminada: {log.filas_importadas} nuevas, {log.filas_actualizadas} actualizadas, {log.filas_fallidas} fallidas.",
    )
    return redirect("rating:importar")


def _valor_inicial(valor):
    if valor.criterio.tipo_dato == "booleano":
        return (
            "si"
            if valor.valor_booleano
            else "no" if valor.valor_booleano is False else ""
        )
    if valor.criterio.tipo_dato == "numero":
        return valor.valor_numero
    return valor.valor_texto


@login_required
def evaluar_empresa(request, empresa_id):
    empresa = get_object_or_404(Empresa, pk=empresa_id)
    anio = timezone.localdate().year
    evaluacion = (
        EvaluacionVisita.objects.filter(empresa=empresa, anio_evaluacion=anio)
        .order_by("-actualizado")
        .first()
    )
    if not evaluacion:
        evaluacion = EvaluacionVisita(
            empresa=empresa,
            evaluado_por=request.user,
            anio_evaluacion=anio,
        )
    return _procesar_evaluacion(request, empresa, evaluacion)


def _datos_iniciales(empresa):
    return {"contactos": [{"nombres": empresa.nombres_apellidos, "apellidos": "", "cargo": empresa.cargo, "telefono": empresa.telefono, "correo1": empresa.email, "correo2": ""}]}


def _leer_respuestas(post):
    respuestas = {}
    for seccion in SECCIONES:
        for question in seccion["questions"]:
            code = question["code"]
            if question["type"] == "repeat":
                columns = {sub["code"]: post.getlist(f"{code}__{sub['code']}[]") for sub in question["subfields"]}
                total = max([len(values) for values in columns.values()] or [0])
                respuestas[code] = [
                    {key: (values[index].strip() if index < len(values) else "") for key, values in columns.items()}
                    for index in range(total)
                    if any(index < len(values) and values[index].strip() for values in columns.values())
                ]
            elif question["type"] == "number_blank" and post.get(f"{code}__blank"):
                respuestas[code] = ""
            else:
                respuestas[code] = post.get(code, "").strip()
    for seccion in SECCIONES:
        for question in seccion["questions"]:
            if question.get("show_if"):
                parent, expected = question["show_if"].split(":", 1)
                if respuestas.get(parent) != expected:
                    respuestas[question["code"]] = [] if question["type"] == "repeat" else ""
    return respuestas


@login_required
def evaluar_visita(request, atencion_id):
    atencion = get_object_or_404(Atencion.objects.select_related("empresa", "responsable"), pk=atencion_id)
    return redirect("rating:evaluar", empresa_id=atencion.empresa_id)


def _procesar_evaluacion(request, empresa, evaluacion):
    perfil, _ = PerfilEvaluacionEmpresa.objects.get_or_create(empresa=empresa, defaults={"datos": _datos_iniciales(empresa)})
    valores = {**perfil.datos, **evaluacion.respuestas}
    redes = valores.get("redes_detalle", [])
    plataformas = {row.get("red") for row in redes if isinstance(row, dict)}
    if valores.get("web_url") and not valores.get("web"):
        valores["web"] = "si"
    for plataforma in ("facebook", "instagram", "linkedin"):
        if plataforma in plataformas and not valores.get(plataforma):
            valores[plataforma] = "si"
    if request.method == "POST":
        respuestas = _leer_respuestas(request.POST)
        puntos, total = puntuar(respuestas)
        evaluacion.empresa = empresa
        evaluacion.evaluado_por = request.user
        evaluacion.anio_evaluacion = timezone.localdate().year
        evaluacion.fecha_evaluacion = timezone.localdate()
        evaluacion.respuestas = respuestas
        evaluacion.puntajes_seccion = {key: str(value) for key, value in puntos.items()}
        evaluacion.puntaje_total = total
        evaluacion.save()
        rating, _ = EmpresaRating.objects.get_or_create(empresa=empresa)
        rating.total = total
        rating.clasificacion = (
            "Alta capacidad" if total >= 21 else "Capacidad media" if total >= 12 else "En desarrollo"
        )
        rating.producto = empresa.oferta_producto_servicio
        rating.save(update_fields=["total", "clasificacion", "producto", "actualizado"])
        perfil.datos = respuestas
        perfil.save(update_fields=["datos", "actualizado"])
        auditar(
            request,
            "editar",
            evaluacion,
            descripcion=f"Evaluación anual {evaluacion.anio_evaluacion} de {empresa.nombre} actualizada",
        )
        messages.success(request, "Evaluación guardada y datos permanentes actualizados.")
        return redirect("rating:evaluar", empresa_id=empresa.pk)
    secciones = preparar_ficha(valores)
    for seccion in secciones:
        seccion["score"] = (evaluacion.puntajes_seccion or {}).get(seccion["code"], "0")
    return render(
        request,
        "rating/ficha.html",
        {
            "empresa": empresa,
            "evaluacion": evaluacion,
            "secciones": secciones,
            "anio": timezone.localdate().year,
        },
    )


@login_required
def exportar_evaluacion(request, evaluacion_id):
    evaluacion = get_object_or_404(
        EvaluacionVisita.objects.select_related("empresa", "atencion", "evaluado_por"),
        pk=evaluacion_id,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Evaluación"
    ws.append(["PROMPERÚ – Ficha de evaluación por visita"])
    ws.append(["Empresa / persona", evaluacion.empresa.nombre])
    ws.append(["Documento", evaluacion.empresa.numero_documento])
    ws.append(["Año de evaluación", evaluacion.anio_evaluacion])
    ws.append(["Fecha de evaluación", evaluacion.fecha_evaluacion])
    ws.append(["Asesor", evaluacion.evaluado_por.get_full_name() if evaluacion.evaluado_por else ""])
    ws.append([])
    ws.append(["Sección", "Pregunta", "Respuesta", "Puntaje de sección"])
    for seccion in SECCIONES:
        for index, question in enumerate(seccion["questions"]):
            value = evaluacion.respuestas.get(question["code"], "")
            if isinstance(value, list):
                value = "; ".join(", ".join(str(v) for v in row.values() if v) for row in value)
            ws.append([seccion["title"], question["label"], value, evaluacion.puntajes_seccion.get(seccion["code"], "") if index == 0 else ""])
    ws.append([])
    ws.append(["PUNTAJE TOTAL", "", "", float(evaluacion.puntaje_total)])
    ws.freeze_panes = "A9"
    for col, width in {"A": 32, "B": 62, "C": 48, "D": 20}.items():
        ws.column_dimensions[col].width = width
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="evaluacion_{evaluacion.empresa_id}_{evaluacion.anio_evaluacion}.xlsx"'
    wb.save(response)
    return response
